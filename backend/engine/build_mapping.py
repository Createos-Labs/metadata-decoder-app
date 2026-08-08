"""
M&A Mapping Engine
==================
Joins catalog exports, contract exports, and royalty statement files into a
master ISRC × rate mapping XLSX.

File bytes are passed in as a dict keyed by detected type. Types are
auto-detected from filenames via `detect_file_type()`.

Main entry points:
  build_mapping(files)        → (xlsx_bytes, stats_dict)
  build_blank_template()      → xlsx_bytes  (no data rows, Instructions tab included)
"""
from __future__ import annotations

import csv
import io
import zipfile
from collections import defaultdict

import openpyxl
import xlrd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Column definitions
# (band, header, width, is_pct, plain_english_description)
# ---------------------------------------------------------------------------

COLUMN_DEFS: list[tuple[str, str, int, bool, str]] = [
    # RECORDING
    ("recording", "ISRC", 16, False,
     "The unique international identifier for this recording — like a barcode for a song. "
     "Every distinct recording (including remixes and live versions) has its own ISRC."),
    ("recording", "Track Title", 28, False,
     "The name of the song/track."),
    ("recording", "Track Version", 16, False,
     "If the track is a remix, radio edit, acoustic version, etc. Blank for the original."),
    ("recording", "UPC / Barcode", 16, False,
     "The barcode of the album or single this track appears on."),
    ("recording", "Release Title", 28, False,
     "The album or single name this track is part of."),
    ("recording", "Cat No", 14, False,
     "The catalog number the label assigned to this release (e.g. GLS016802)."),
    ("recording", "Label", 18, False,
     "The label that released this recording."),
    ("recording", "P Line", 20, False,
     "The phonographic copyright line (℗ Year Label Name). Confirms who owns the master recording."),
    ("recording", "Release Date", 14, False,
     "The commercial release date of the product this track appears on."),
    ("recording", "Duration", 10, False,
     "Track length in minutes:seconds."),
    ("recording", "Format", 14, False,
     "The product format — CD, LP, Digital, etc."),
    ("recording", "Distribution Channel", 22, False,
     "The channel through which this product was distributed (e.g. Digital, Physical)."),
    ("recording", "Price Category", 16, False,
     "The pricing tier (e.g. Full Price, Mid Price, Budget)."),
    ("recording", "Catalogue Group", 20, False,
     "Internal grouping in the royalty system used to apply specific rates to groups of releases. "
     "Populate manually from contract records."),
    ("recording", "Dealer Price", 14, False,
     "The wholesale/dealer price for physical products in USD."),
    # ARTIST
    ("artist", "Artist Name", 24, False,
     "The artist or rights holder this row's contract is associated with."),
    ("artist", "Track Sales Contract Name", 28, False,
     "The name of the sales royalty contract that covers this track for this artist."),
    ("artist", "Track Sales Contract %", 22, True,
     "The percentage share of sales income allocated to this artist (proration). "
     "Relevant when multiple artists share a contract."),
    ("artist", "Track Costs Contract Name", 28, False,
     "The name of the costs contract covering this track. Costs (recording, marketing) can "
     "be tracked under a separate contract linked to the sales contract."),
    ("artist", "Track Costs Contract %", 22, True,
     "The percentage share of costs this artist is responsible for."),
    # CONTRACT
    ("contract", "Contract Name", 28, False,
     "The full name of the royalty contract in the system."),
    ("contract", "Payee Name", 22, False,
     "The person or entity who receives royalty payments under this contract."),
    ("contract", "Contract Type", 18, False,
     "The deal structure: Royalty (% of revenue), Profit Share (% of net profit), or License (flat fee/buyout)."),
    ("contract", "Accounting Period", 18, False,
     "How often royalty statements are issued — Monthly, Quarterly, Half-Yearly, or Yearly. "
     "Populate manually."),
    ("contract", "Currency", 12, False,
     "The currency royalties are paid in."),
    ("contract", "Profit Share %", 16, True,
     "For Profit Share deals only — the percentage of net profit allocated to this artist."),
    ("contract", "Contract Start Date", 18, False,
     "The date the contract term begins."),
    ("contract", "Contract End Date", 18, False,
     "The date the contract term ends. Blank means perpetual or no fixed end date."),
    ("contract", "Notes", 30, False,
     "Any additional notes or flags from the contract record."),
    # RATE
    ("rate", "Rate Territory", 16, False,
     "The geographic territory this rate row applies to. ISO2 codes (US, CA) or WW for Worldwide."),
    ("rate", "Rate Channel", 18, False,
     "The distribution channel this rate applies to (e.g. All Physical, All Digital, Streaming)."),
    ("rate", "Rate Configuration", 20, False,
     "The release format this rate applies to (e.g. Album, Single, All)."),
    ("rate", "Rate Price Category", 18, False,
     "The pricing tier this rate applies to (e.g. Full Price, Mid Price, Budget)."),
    ("rate", "Rate Source", 18, False,
     "The specific income source this rate applies to (e.g. iTunes/Apple, Spotify, All Digital)."),
    ("rate", "Rate Type / Basis", 20, False,
     "What the rate is calculated against — the revenue base (e.g. % Wholesale, % Net Receipts)."),
    ("rate", "Rate %", 10, True,
     "The actual royalty rate as a percentage applied against the Rate Basis."),
    ("rate", "Reserve %", 12, True,
     "Percentage of earned royalties held back as a reserve against returns or chargebacks."),
    ("rate", "Reserve Basis", 18, False,
     "Whether the reserve is calculated before or after other deductions. Populate manually."),
    ("rate", "Esc. Threshold Type", 20, False,
     "For escalating rates — the unit used to measure when escalation triggers. Blank if no escalation."),
    ("rate", "Esc. When", 14, False,
     "Whether the escalated rate applies to all sales (Cumulative) or only sales above the threshold (Marginal)."),
    ("rate", "Esc. Threshold", 16, False,
     "The number or revenue amount at which the escalation triggers."),
    ("rate", "Esc. Rate %", 12, True,
     "The higher royalty rate that applies once the escalation threshold is met."),
    # FINANCIALS
    ("financials", "Balance (Latest Period)", 22, False,
     "The artist's royalty balance from their most recent available statement, with the period in brackets. "
     "ISRC-level from Sales CSVs for Profit Share artists; contract-level Ending Balance for all others."),
    ("financials", "Advance Remaining", 18, False,
     "The unrecouped advance still outstanding — how much the artist still owes before receiving payments. "
     "Populate from contract or statement records."),
    ("financials", "Min Payout", 14, False,
     "The minimum balance required before a royalty payment is triggered."),
    ("financials", "Deduct Withholding Tax", 22, False,
     "Whether withholding tax is deducted from royalties before payment. Relevant for international artists."),
    ("financials", "Cross Contract Name", 26, False,
     "If cross-collateralized, the name of the linked contract. Cross-collateralization means the artist "
     "won't receive royalties until all advances across both contracts are recouped."),
    ("financials", "Cross Contract Process", 22, False,
     "How funds are transferred between cross-collateralized contracts."),
    ("financials", "Cross Contract Link", 26, False,
     "The specific contract name this one transfers funds to/from."),
    ("financials", "Cross Contract %", 18, True,
     "The percentage of income transferred between cross-collateralized contracts."),
]

# ---------------------------------------------------------------------------
# Styling constants
# ---------------------------------------------------------------------------

BANDS: dict[str, dict[str, str]] = {
    "recording":  {"header": "B5696A", "rowA": "F7E8E8", "rowB": "EDCFCF"},
    "artist":     {"header": "8B7BAB", "rowA": "EEE8F5", "rowB": "DDD0F0"},
    "contract":   {"header": "6A9178", "rowA": "E8F2EC", "rowB": "CFE8D8"},
    "rate":       {"header": "B8976A", "rowA": "F7F0E4", "rowB": "EDE0C8"},
    "financials": {"header": "957A8F", "rowA": "F2EBF5", "rowB": "E2D0E8"},
}

BAND_LABELS = {
    "recording": "RECORDING",
    "artist": "ARTIST",
    "contract": "CONTRACT",
    "rate": "RATE",
    "financials": "FINANCIALS",
}

PCT_FORMAT = '0.00"%"'


# ---------------------------------------------------------------------------
# File type detection
# ---------------------------------------------------------------------------

def detect_file_type(filename: str) -> str:
    """
    Returns one of: 'catalog', 'contract_terms', 'isrc_links', 'payees',
                    'statement_zip', 'statement_csv', 'statement_xlsx', 'unknown'
    Based on Glassnote export filename conventions.
    """
    n = filename.lower()
    if n.endswith(".zip"):
        return "statement_zip"
    if "walbum" in n or "w_album" in n or "walbums" in n:
        return "isrc_links"
    if "terms" in n and "contract" in n:
        return "contract_terms"
    if "payee" in n:
        return "payees"
    if "product" in n or ("track" in n and "contract" not in n):
        return "catalog"
    if n.endswith(".csv"):
        return "statement_csv"
    if n.endswith(".xlsx") and ("fullreport" in n or "revenue_details" in n):
        return "statement_xlsx"
    if "remaining term" in n or "contracts and remaining" in n:
        return "orchard_contracts"
    if "connection advance" in n or ("advance" in n and "balance" in n):
        return "advance_balances"
    return "unknown"


# ---------------------------------------------------------------------------
# Data loaders
# ---------------------------------------------------------------------------

def _load_xlsx(data: bytes) -> tuple[openpyxl.Workbook, object]:
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    ws = wb.active
    return wb, ws


def load_catalog(data: bytes) -> dict[str, dict]:
    """Returns {isrc: {field: value, ...}}"""
    wanted = [
        "isrc", "track-title", "primary-track-artist", "upc", "album-title",
        "catalog-no", "label-name", "p-line", "release-date",
        "track-minutes", "track-seconds", "product-type", "full-price (usd)",
    ]
    try:
        wb = xlrd.open_workbook(file_contents=data)
        ws = wb.sheets()[0]
        hdrs = {ws.cell_value(0, c): c for c in range(ws.ncols)}
        result: dict[str, dict] = {}
        for r in range(1, ws.nrows):
            isrc_col = hdrs.get("isrc", -1)
            isrc = ws.cell_value(r, isrc_col) if isrc_col >= 0 else None
            if not isrc:
                continue
            rec: dict = {}
            for k in wanted:
                col = hdrs.get(k, -1)
                v = ws.cell_value(r, col) if col >= 0 else ""
                rec[k] = v
            result[str(isrc).strip()] = rec
        return result
    except Exception:
        pass
    _, ws = _load_xlsx(data)
    hdrs = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
    result = {}
    for r in range(2, ws.max_row + 1):
        isrc = ws.cell(r, hdrs.get("isrc", 0)).value if "isrc" in hdrs else None
        if not isrc:
            continue
        rec = {}
        for k in wanted:
            col = hdrs.get(k)
            rec[k] = ws.cell(r, col).value if col else ""
        result[str(isrc).strip()] = rec
    return result


def load_contract_terms(data: bytes) -> dict[str, list[dict]]:
    """Returns {contract_title: [term_dict, ...]}"""
    wanted = [
        "contract-title", "payee", "region", "channel", "source", "price",
        "rate-type", "rate", "reserve-rate", "term-start", "term-end", "comments",
    ]
    try:
        wb = xlrd.open_workbook(file_contents=data)
        ws = wb.sheets()[0]
        hdrs = {ws.cell_value(0, c): c for c in range(ws.ncols)}
        result: dict[str, list[dict]] = defaultdict(list)
        title_col = hdrs.get("contract-title", -1)
        if title_col >= 0:
            for r in range(1, ws.nrows):
                title = ws.cell_value(r, title_col)
                if not title:
                    continue
                term: dict = {}
                for k in wanted:
                    col = hdrs.get(k, -1)
                    v = ws.cell_value(r, col) if col >= 0 else ""
                    # xlrd returns dates as floats; convert to string
                    if ws.cell_type(r, col) == xlrd.XL_CELL_DATE if col >= 0 else False:
                        try:
                            dt = xlrd.xldate_as_datetime(v, wb.datemode)
                            v = dt.strftime("%Y-%m-%d")
                        except Exception:
                            pass
                    term[k] = v or ""
                result[str(title).strip()].append(term)
        return dict(result)
    except Exception:
        pass
    _, ws = _load_xlsx(data)
    hdrs = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
    result = defaultdict(list)
    for r in range(2, ws.max_row + 1):
        title_col = hdrs.get("contract-title")
        if not title_col:
            break
        title = ws.cell(r, title_col).value
        if not title:
            continue
        term = {}
        for k in wanted:
            col = hdrs.get(k)
            v = ws.cell(r, col).value if col else ""
            if hasattr(v, "strftime"):
                v = v.strftime("%Y-%m-%d")
            term[k] = v or ""
        result[str(title).strip()].append(term)
    return dict(result)


def load_isrc_links(data: bytes) -> list[dict]:
    """Returns [{isrc, contract-title, artist, payee, proration, cross-collateralize, ...}]"""
    try:
        wb = xlrd.open_workbook(file_contents=data)
    except Exception:
        # Fallback: try as xlsx
        _, ws = _load_xlsx(data)
        return _load_isrc_links_from_ws(ws)
    ws = wb.sheets()[0]
    hdrs = {ws.cell_value(0, c): c for c in range(ws.ncols)}
    wanted = [
        "track-isrc", "contract-title", "artist", "payee",
        "track-title", "album-title", "catalog-no", "proration", "cross-collateralize",
    ]
    result = []
    for r in range(1, ws.nrows):
        isrc = ws.cell_value(r, hdrs.get("track-isrc", -1)) if "track-isrc" in hdrs else None
        if not isrc:
            continue
        rec: dict = {}
        for k in wanted:
            col = hdrs.get(k, -1)
            rec[k] = ws.cell_value(r, col) if col >= 0 else ""
        result.append(rec)
    return result


def _load_isrc_links_from_ws(ws) -> list[dict]:
    hdrs = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
    wanted = [
        "track-isrc", "contract-title", "artist", "payee",
        "track-title", "album-title", "catalog-no", "proration", "cross-collateralize",
    ]
    result = []
    for r in range(2, ws.max_row + 1):
        isrc_col = hdrs.get("track-isrc")
        if not isrc_col:
            break
        isrc = ws.cell(r, isrc_col).value
        if not isrc:
            continue
        rec: dict = {}
        for k in wanted:
            col = hdrs.get(k)
            rec[k] = ws.cell(r, col).value if col else ""
        result.append(rec)
    return result


def load_statement_csv(data: bytes, net_by: dict) -> int:
    """
    Parse a single Profit Share Sales Export CSV into net_by.
    Streams row-by-row to avoid holding the full decoded text in memory.
    Returns row count.
    """
    try:
        # Use a raw bytes buffer so we don't decode the entire file at once
        buf = io.TextIOWrapper(io.BytesIO(data), encoding="utf-8-sig")
        reader = csv.DictReader(buf)
        count = 0
        for row in reader:
            contract = row.get("Contract Name", row.get("﻿Contract Name", "")).strip()
            isrc = row.get("ISRC", "").strip() or "__release__"
            # Sale Date format is YYYY-MM (first 7 chars of whatever date string)
            period = (row.get("Sale Date") or "").strip()[:7]
            try:
                net = float(row.get("Net Payable") or 0)
            except (ValueError, TypeError):
                net = 0.0
            if contract and period:
                net_by[contract][isrc][period] += net
                count += 1
        return count
    except Exception:
        return 0


def _extract_period_from_filename(filename: str) -> str | None:
    """Extract a period key (e.g. '1h25') from an Orchard monthly XLSX filename."""
    import re
    n = filename.lower()
    # Try "YYYYMMDD_MonYYYY_" prefix — e.g. "20230222_Jan2023_"
    m = re.search(r"_([a-z]{3})(\d{4})_", n)
    if m:
        mon_abbr = m.group(1)
        year = m.group(2)[-2:]
        months_h1 = {"jan", "feb", "mar", "apr", "may", "jun"}
        half = "1h" if mon_abbr in months_h1 else "2h"
        return f"{half}{year}"
    # Try "_Month_YYYY_" pattern — e.g. "_February_2026_"
    months = {
        "january": "1h", "february": "1h", "march": "1h",
        "april": "1h", "may": "1h", "june": "1h",
        "july": "2h", "august": "2h", "september": "2h",
        "october": "2h", "november": "2h", "december": "2h",
    }
    m2 = re.search(r"_([a-z]+)_(\d{4})_", n)
    if m2:
        mon = m2.group(1)
        year = m2.group(2)[-2:]
        half = months.get(mon)
        if half:
            return f"{half}{year}"
    return None


def load_statement_xlsx(data: bytes, filename: str, best_balance: dict) -> int:
    """
    Parse an Orchard monthly fullreport XLSX for Net Payable data grouped by contract.
    Uses read_only + streaming to avoid loading the full file into memory.
    Mutates best_balance in place. Returns number of data rows processed.
    """
    period = _extract_period_from_filename(filename)
    if not period:
        return 0
    try:
        wb = openpyxl.load_workbook(
            io.BytesIO(data), data_only=True, read_only=True
        )
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        header = next(rows_iter, None)
        if not header:
            wb.close()
            return 0
        hdrs = {
            str(v).strip().lower(): i
            for i, v in enumerate(header)
            if v is not None
        }
        contract_col = hdrs.get("contract name")
        net_col = hdrs.get("net payable")
        if contract_col is None or net_col is None:
            wb.close()
            return 0
        totals: dict[str, float] = defaultdict(float)
        count = 0
        for row in rows_iter:
            contract = str(row[contract_col] or "").strip()
            if not contract:
                continue
            try:
                net = float(row[net_col] or 0)
            except (ValueError, TypeError):
                net = 0.0
            totals[contract] += net
            count += 1
        wb.close()
        for contract, total in totals.items():
            key = contract.lower()
            existing = best_balance.get(key)
            if existing is None or _period_rank(period) > _period_rank(
                existing["period"]
            ):
                best_balance[key] = {
                    "balance": round(total, 2),
                    "period": period,
                    "artist": contract,
                }
        return count
    except Exception:
        return 0


def load_orchard_contracts(data: bytes) -> dict:
    """
    Parse 'Contracts and Remaining Term' summary XLSX.
    Returns {artist_name_lower: {type, contract_date, territory, royalty_split,
                                  profit_share_pct, expiration_date, notes}}
    """
    import re as _re
    result: dict = {}
    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True, read_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        header = next(rows_iter, None)
        if not header:
            wb.close()
            return result
        hdrs = {str(v).strip().lower(): i for i, v in enumerate(header) if v is not None}
        skip_values = {"future revenue stream", "catalog", "connection", "total", "artist", ""}
        for row in rows_iter:
            artist = str(row[hdrs.get("artist", 0)] or "").strip()
            if not artist or artist.lower() in skip_values:
                continue
            # Parse royalty_split like "50% Net Proceeds" → 50.0
            royalty_raw = str(row[hdrs.get("royalty split", -1)] or "") if hdrs.get("royalty split") is not None else ""
            pct_match = _re.search(r"(\d+(?:\.\d+)?)\s*%", royalty_raw)
            profit_share_pct = float(pct_match.group(1)) if pct_match else ""

            def _date(col_name: str):
                idx = hdrs.get(col_name)
                if idx is None:
                    return ""
                v = row[idx]
                if v is None:
                    return ""
                if hasattr(v, "strftime"):
                    return v.strftime("%Y-%m-%d")
                return str(v).strip()

            result[artist.lower()] = {
                "artist": artist,
                "type": str(row[hdrs.get("type", -1)] or "").strip() if hdrs.get("type") is not None else "",
                "contract_date": _date("contract date"),
                "territory": str(row[hdrs.get("territory", -1)] or "").strip() if hdrs.get("territory") is not None else "",
                "royalty_split": royalty_raw.strip(),
                "profit_share_pct": profit_share_pct,
                "expiration_date": _date("licensing period expiration") or _date("licence period experation") or _date("trigger date"),
                "notes": str(row[hdrs.get("notes:", -1)] or "").strip() if hdrs.get("notes:") is not None else "",
            }
        wb.close()
    except Exception:
        pass
    return result


def load_advance_balances(data: bytes) -> dict:
    """
    Parse a simple two-column advance balances XLSX (Name | Amount).
    Skips header/title rows. Returns {name_lower: balance_float}.
    """
    result: dict = {}
    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True, read_only=True)
        ws = wb.active
        for row in ws.iter_rows(values_only=True):
            name = row[0] if row else None
            val = row[1] if len(row) > 1 else None
            if not name or val is None:
                continue
            name_str = str(name).strip()
            if not name_str:
                continue
            try:
                balance = float(val)
            except (ValueError, TypeError):
                continue
            result[name_str.lower()] = {"name": name_str, "balance": balance}
        wb.close()
    except Exception:
        pass
    return result


def _period_rank(p: str) -> int:
    return {"1h23": 1, "2h23": 2, "1h24": 3, "2h24": 4, "1h25": 5}.get(
        str(p).lower().strip(), 0
    )


def load_statement_zip(
    data: bytes,
    net_by: dict,
    best_balance: dict,
) -> int:
    """
    Parse a statement zip. Mutates net_by and best_balance in place.
    Returns number of files processed.
    """
    processed = 0
    try:
        with zipfile.ZipFile(io.BytesIO(data)) as z:
            for entry in z.infolist():
                fn = entry.filename.lower()
                if entry.file_size == 0:
                    continue
                # Sales CSV
                if "sales" in fn and fn.endswith(".csv"):
                    try:
                        with z.open(entry.filename) as f:
                            reader = csv.DictReader(
                                io.TextIOWrapper(f, encoding="utf-8-sig")
                            )
                            for row in reader:
                                contract = (
                                    row.get("Contract Name", row.get("﻿Contract Name", ""))
                                    .strip()
                                )
                                isrc = row.get("ISRC", "").strip() or "__release__"
                                period = row.get("Sale Date", "").strip()[:7]
                                try:
                                    net = float(row.get("Net Payable", 0) or 0)
                                except (ValueError, TypeError):
                                    net = 0.0
                                if contract and period:
                                    net_by[contract][isrc][period] += net
                            processed += 1
                    except Exception:
                        pass
                # Statement XLSX
                elif fn.endswith(".xlsx") and entry.file_size > 5000:
                    period = next(
                        (p for p in ["1h25", "2h24", "1h24", "2h23", "1h23"]
                         if p in fn),
                        None,
                    )
                    if not period:
                        continue
                    try:
                        with z.open(entry.filename) as f:
                            wb2 = openpyxl.load_workbook(
                                io.BytesIO(f.read()), data_only=True
                            )
                            if "Statement" not in wb2.sheetnames:
                                continue
                            ws2 = wb2["Statement"]
                            artist = None
                            balance = None
                            for r in range(1, min(25, ws2.max_row + 1)):
                                v0 = ws2.cell(r, 1).value
                                v1 = ws2.cell(r, 2).value
                                if r == 7 and v0:
                                    artist = str(v0).strip()
                                if v0 == "Ending Balance:" and v1 is not None:
                                    try:
                                        balance = float(v1)
                                    except (ValueError, TypeError):
                                        pass
                            if artist and balance is not None:
                                key = artist.lower()
                                existing = best_balance.get(key)
                                if existing is None or _period_rank(period) > _period_rank(
                                    existing["period"]
                                ):
                                    best_balance[key] = {
                                        "balance": balance,
                                        "period": period,
                                        "artist": artist,
                                    }
                            processed += 1
                    except Exception:
                        pass
    except Exception:
        pass
    return processed


# ---------------------------------------------------------------------------
# Row builder
# ---------------------------------------------------------------------------

_BLANK_TERM: dict = {
    k: ""
    for k in [
        "payee", "region", "channel", "source", "price",
        "rate-type", "rate", "reserve-rate", "term-start", "term-end", "comments",
    ]
}


def _build_data_rows(
    catalog: dict,
    contract_terms: dict,
    links: list[dict],
    net_by: dict,
    best_balance: dict,
    orchard_contracts: dict | None = None,
    advance_balances: dict | None = None,
) -> list[tuple]:
    def latest_net(contract: str, isrc: str):
        key = isrc or "__release__"
        p = net_by.get(contract, {}).get(key, {})
        if not p:
            p = net_by.get(contract, {}).get("__release__", {})
        if not p:
            return "", ""
        latest = max(p.keys())
        return round(p[latest], 2), latest

    def fallback_balance(payee: str):
        if not payee:
            return "", ""
        v = best_balance.get(str(payee).strip().lower())
        if v:
            return v["balance"], v["period"].upper()
        return "", ""

    rows = []
    for link in links:
        isrc = str(link.get("track-isrc", "") or "").strip()
        contract = str(link.get("contract-title", "") or "").strip()
        if not isrc or not contract:
            continue
        track = catalog.get(isrc, {})
        terms = contract_terms.get(contract, [_BLANK_TERM])

        mins = track.get("track-minutes", "")
        secs = track.get("track-seconds", "")
        try:
            duration = f"{int(mins)}:{int(secs):02d}" if mins != "" else ""
        except (ValueError, TypeError):
            duration = ""

        rd = track.get("release-date", "")
        if hasattr(rd, "strftime"):
            rd = rd.strftime("%Y-%m-%d")

        for term in terms:
            td = term.get("term-start", "")
            te = term.get("term-end", "")
            if hasattr(td, "strftime"):
                td = td.strftime("%Y-%m-%d")
            if hasattr(te, "strftime"):
                te = te.strftime("%Y-%m-%d")

            payee = str(term.get("payee", "") or link.get("payee", "") or "").strip()
            proration = link.get("proration", "")

            net_amt, period = latest_net(contract, isrc)
            if net_amt != "":
                bal_str = f"{net_amt}  ({period})"
            else:
                fb, fp = fallback_balance(payee)
                bal_str = f"{fb}  ({fp})" if fb != "" else ""

            # Orchard contract summary — look up by artist name (case-insensitive)
            artist_name = str(link.get("artist", "") or track.get("primary-track-artist", "") or "").strip()
            oc: dict = {}
            if orchard_contracts:
                oc = orchard_contracts.get(artist_name.lower(), {})
                if not oc:
                    # Partial match: artist key starts with or is contained in artist_name
                    for k, v in orchard_contracts.items():
                        if artist_name.lower().startswith(k) or k.startswith(artist_name.lower()):
                            oc = v
                            break

            # Advance balance — look up by payee name (case-insensitive)
            advance_str = ""
            if advance_balances and payee:
                ab = advance_balances.get(payee.lower(), {})
                if ab:
                    advance_str = str(ab.get("balance", ""))

            # Prefer terms-export dates; fall back to orchard contract dates
            contract_start = td or oc.get("contract_date", "")
            contract_end = te or oc.get("expiration_date", "")
            territory = term.get("region", "") or oc.get("territory", "")
            contract_type = term.get("rate-type", "") or oc.get("type", "")
            profit_share = oc.get("profit_share_pct", "") if oc else ""
            notes = term.get("comments", "") or oc.get("notes", "")

            rows.append((
                isrc,
                track.get("track-title", "") or link.get("track-title", ""),
                "",  # Track Version
                track.get("upc", ""),
                track.get("album-title", "") or link.get("album-title", ""),
                track.get("catalog-no", "") or link.get("catalog-no", ""),
                track.get("label-name", ""),
                track.get("p-line", ""),
                rd,
                duration,
                track.get("product-type", ""),
                "",  # Distribution Channel
                term.get("price", ""),
                "",  # Catalogue Group
                track.get("full-price (usd)", ""),
                artist_name,
                contract,
                proration,  # Track Sales Contract %
                contract,   # Track Costs Contract Name
                proration,  # Track Costs Contract %
                contract,
                payee,
                contract_type,
                "",  # Accounting Period
                "USD",
                profit_share,
                contract_start,
                contract_end,
                notes,
                territory,
                term.get("channel", ""),
                "",  # Rate Configuration
                term.get("price", ""),
                term.get("source", ""),
                term.get("rate-type", ""),
                term.get("rate", ""),
                term.get("reserve-rate", ""),
                "",  # Reserve Basis
                "",  # Esc. Threshold Type
                "",  # Esc. When
                "",  # Esc. Threshold
                "",  # Esc. Rate %
                bal_str,
                advance_str,
                "",  # Min Payout
                "",  # Deduct Withholding Tax
                link.get("cross-collateralize", ""),
                "",  # Cross Contract Process
                "",  # Cross Contract Link
                "",  # Cross Contract %
            ))
    return rows


# ---------------------------------------------------------------------------
# Workbook builders
# ---------------------------------------------------------------------------

def _make_styles(wb):
    fills = {
        (b, k): PatternFill("solid", fgColor=BANDS[b][k])
        for b in BANDS
        for k in ("header", "rowA", "rowB")
    }
    hdr_font = Font(name="Calibri", bold=True, color="FFFFFF", size=9)
    band_font = Font(name="Calibri", bold=True, color="FFFFFF", size=10, italic=True)
    body_font = Font(name="Calibri", size=9, color="3A2A2A")
    mid_align = Alignment(vertical="center")
    ctr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    divider = Border(bottom=Side(style="medium", color="B5696A"))
    hairline = Border(bottom=Side(style="hair", color="E0D0D0"))
    return fills, hdr_font, band_font, body_font, mid_align, ctr_align, divider, hairline


def _build_mapping_sheet(ws, rows: list[tuple]):
    fills, hdr_font, band_font, body_font, mid_align, ctr_align, divider, hairline = (
        _make_styles(ws.parent)
    )

    # Band ranges (row 1)
    band_ranges: dict[str, list[int]] = {}
    for i, (band, *_) in enumerate(COLUMN_DEFS, 1):
        if band not in band_ranges:
            band_ranges[band] = [i, i]
        else:
            band_ranges[band][1] = i

    for band, (s, e) in band_ranges.items():
        if s != e:
            ws.merge_cells(start_row=1, start_column=s, end_row=1, end_column=e)
        c = ws.cell(1, s)
        c.value = BAND_LABELS[band]
        c.fill = fills[(band, "header")]
        c.font = band_font
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    # Column headers (row 2)
    pct_cols: set[int] = set()
    for i, (band, header, width, is_pct, *_) in enumerate(COLUMN_DEFS, 1):
        c = ws.cell(2, i)
        c.value = header
        c.fill = fills[(band, "header")]
        c.font = hdr_font
        c.alignment = ctr_align
        ws.column_dimensions[get_column_letter(i)].width = width
        ws.row_dimensions[2].height = 36
        if is_pct:
            pct_cols.add(i)

    if not rows:
        return

    # Alternating ISRC colors
    isrc_order: list[str] = []
    seen: set[str] = set()
    for r in rows:
        if r[0] not in seen:
            seen.add(r[0])
            isrc_order.append(r[0])
    isrc_toggle = {isrc: (i % 2 == 0) for i, isrc in enumerate(isrc_order)}

    cur = 3
    for idx, row in enumerate(rows):
        isrc = row[0]
        is_last = (idx == len(rows) - 1) or (rows[idx + 1][0] != isrc)
        toggle = isrc_toggle.get(isrc, True)

        for ci, (band, *_) in enumerate(COLUMN_DEFS, 1):
            c = ws.cell(cur, ci)
            val = row[ci - 1]
            if val == "" or val is None:
                c.value = None
            elif ci in pct_cols:
                try:
                    c.value = float(val)
                    c.number_format = PCT_FORMAT
                except (ValueError, TypeError):
                    c.value = val
            else:
                c.value = val
            c.fill = fills[(band, "rowA" if toggle else "rowB")]
            c.font = body_font
            c.alignment = mid_align
            c.border = divider if is_last else hairline
        ws.row_dimensions[cur].height = 16
        cur += 1

    ws.freeze_panes = "C3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(COLUMN_DEFS))}2"


def _build_instructions_sheet(ws):
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 70
    ws.column_dimensions["D"].width = 22

    TITLE_FONT = Font(name="Calibri", bold=True, size=18, color="1A1612")
    H1_FILL = PatternFill("solid", fgColor="B5696A")
    H1_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    BODY_FONT = Font(name="Calibri", size=10, color="1A1612")
    MUTED_FONT = Font(name="Calibri", size=9, color="666666", italic=True)
    LIGHT_FILL = PatternFill("solid", fgColor="FAF6F4")
    WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")
    TOP_ALIGN = Alignment(vertical="top", wrap_text=True)
    MID_ALIGN = Alignment(vertical="center", wrap_text=True)

    def heading(r, text):
        ws.merge_cells(f"B{r}:D{r}")
        c = ws.cell(r, 2)
        c.value = text
        c.font = H1_FONT
        c.fill = H1_FILL
        c.alignment = MID_ALIGN
        ws.row_dimensions[r].height = 24
        return r + 1

    def body(r, text, height=None):
        ws.merge_cells(f"B{r}:D{r}")
        c = ws.cell(r, 2)
        c.value = text
        c.font = BODY_FONT
        c.fill = LIGHT_FILL
        c.alignment = TOP_ALIGN
        ws.row_dimensions[r].height = height or 20
        return r + 1

    r = 1
    # Title
    ws.merge_cells(f"B{r}:D{r}")
    c = ws.cell(r, 2)
    c.value = "M&A Mapping Template — Instructions & Field Guide"
    c.font = TITLE_FONT
    c.alignment = Alignment(vertical="center")
    ws.row_dimensions[r].height = 36
    r += 1

    ws.merge_cells(f"B{r}:D{r}")
    c = ws.cell(r, 2)
    c.value = "Master mapping table — one row per ISRC × rate term"
    c.font = MUTED_FONT
    ws.row_dimensions[r].height = 18
    r += 2

    # What is this?
    r = heading(r, "WHAT IS THIS FILE?")
    r = body(r,
        "This is the master mapping table for an acquisition. It joins every ISRC in the seller's "
        "catalog with the royalty contracts and rate terms that cover it. Each row represents one "
        "ISRC × one rate term — so a single track may appear many times if it has multiple rate "
        "terms (e.g. different rates by territory or channel). The purpose is to verify that every "
        "ISRC is correctly mapped to a contract and rate before ingestion into Label Engine (LE).",
        height=72,
    )
    r += 1

    # How to read
    r = heading(r, "HOW TO READ THE MAPPING TEMPLATE TAB")
    band_info = [
        ("recording",  "B5696A", "Blush / Rose",   "Recording metadata: what the track is, where it lives, and how it's packaged."),
        ("artist",     "8B7BAB", "Lavender",        "Artist and contract assignment: who the track belongs to and which contract covers it."),
        ("contract",   "6A9178", "Sage Green",      "Contract-level terms: deal type, dates, currency, and payee."),
        ("rate",       "B8976A", "Champagne",       "Rate terms: the actual royalty rates, territories, channels, and escalation clauses."),
        ("financials", "957A8F", "Mauve",           "Financial snapshot: most recent balance from statements plus cross-contract info."),
    ]
    for band, color, color_name, desc in band_info:
        ws.cell(r, 2).value = f"  {band.upper()}"
        ws.cell(r, 2).font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
        ws.cell(r, 2).fill = PatternFill("solid", fgColor=color)
        ws.cell(r, 2).alignment = MID_ALIGN
        ws.merge_cells(f"C{r}:D{r}")
        ws.cell(r, 3).value = f"{color_name}: {desc}"
        ws.cell(r, 3).font = BODY_FONT
        ws.cell(r, 3).fill = LIGHT_FILL
        ws.cell(r, 3).alignment = MID_ALIGN
        ws.row_dimensions[r].height = 24
        r += 1
    r += 1

    # Tips
    r = heading(r, "NAVIGATION TIPS")
    for tip in [
        "• Columns A–B (ISRC and Track Title) are frozen — scroll right and they stay visible.",
        "• A rose/pink divider line separates each ISRC group. Alternating row colors also help.",
        "• Use the auto-filter on row 2 to filter by artist, territory, channel, contract, or any other field.",
        "• % columns are formatted as numbers with a % sign — '25.00%' means 25%, not 0.25.",
        "• 'Balance (Latest Period)' shows the most recent balance with period in brackets.",
        "• Blank cells in the Financials band mean no statement data was available for that artist.",
    ]:
        r = body(r, tip)
    r += 1

    # Column guide
    r = heading(r, "COLUMN-BY-COLUMN GUIDE")
    # Table header
    for ci, lbl in [(2, "Column"), (3, "What it means"), (4, "Band")]:
        c = ws.cell(r, ci)
        c.value = lbl
        c.font = Font(name="Calibri", bold=True, size=9, color="666666")
        c.fill = PatternFill("solid", fgColor="F0EAE8")
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[r].height = 18
    r += 1

    alt = True
    prev_band = None
    for (band, header, width, is_pct, desc) in COLUMN_DEFS:
        if band != prev_band:
            ws.merge_cells(f"B{r}:D{r}")
            c = ws.cell(r, 2)
            c.value = f"  {BAND_LABELS[band]}"
            c.font = Font(name="Calibri", bold=True, size=9, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor=BANDS[band]["header"])
            c.alignment = MID_ALIGN
            ws.row_dimensions[r].height = 16
            r += 1
            prev_band = band
            alt = True

        row_fill = LIGHT_FILL if alt else WHITE_FILL
        c = ws.cell(r, 2)
        c.value = header + (" *%*" if is_pct else "")
        c.font = Font(name="Courier New", size=9, color=BANDS[band]["header"], bold=True)
        c.fill = row_fill
        c.alignment = TOP_ALIGN

        c2 = ws.cell(r, 3)
        c2.value = desc
        c2.font = Font(name="Calibri", size=9, color="1A1612")
        c2.fill = row_fill
        c2.alignment = TOP_ALIGN

        c3 = ws.cell(r, 4)
        c3.value = band.upper()
        c3.font = Font(name="Calibri", size=9, color="FFFFFF", bold=True)
        c3.fill = PatternFill("solid", fgColor=BANDS[band]["header"])
        c3.alignment = Alignment(horizontal="center", vertical="center")

        ws.row_dimensions[r].height = 40
        r += 1
        alt = not alt

    r += 1
    ws.merge_cells(f"B{r}:D{r}")
    c = ws.cell(r, 2)
    c.value = "* Columns marked *%* are formatted as percentages (e.g. 25.00% means 25%, not 0.25)."
    c.font = MUTED_FONT
    ws.row_dimensions[r].height = 16


def _build_source_guide_sheet(ws, stats: dict):
    """Generated after the tool runs — summarizes what was found."""
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 42
    ws.column_dimensions["D"].width = 32

    H1_FILL = PatternFill("solid", fgColor="6A9178")
    H1_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    BODY_FONT = Font(name="Calibri", size=10, color="1A1612")
    MUTED_FONT = Font(name="Calibri", size=9, color="666666", italic=True)
    LIGHT_FILL = PatternFill("solid", fgColor="F4F9F6")
    MID_ALIGN = Alignment(vertical="center", wrap_text=True)
    TOP_ALIGN = Alignment(vertical="top", wrap_text=True)

    r = 1
    ws.merge_cells(f"B{r}:D{r}")
    c = ws.cell(r, 2)
    c.value = "Source Guide — Where each column's data comes from"
    c.font = Font(name="Calibri", bold=True, size=16, color="1A1612")
    c.alignment = Alignment(vertical="center")
    ws.row_dimensions[r].height = 32
    r += 1

    ws.merge_cells(f"B{r}:D{r}")
    c = ws.cell(r, 2)
    c.value = "This tab lists the source file and original field name for every column in the Mapping Template."
    c.font = MUTED_FONT
    ws.row_dimensions[r].height = 18
    r += 2

    # Generation stats
    ws.merge_cells(f"B{r}:D{r}")
    c = ws.cell(r, 2)
    c.value = "GENERATION SUMMARY"
    c.font = H1_FONT
    c.fill = H1_FILL
    c.alignment = MID_ALIGN
    ws.row_dimensions[r].height = 24
    r += 1

    for label, value in [
        ("Total ISRCs", stats.get("isrc_count", "—")),
        ("Data rows (ISRC × rate term)", stats.get("row_count", "—")),
        ("Artists with balance data", stats.get("artists_with_balance", "—")),
        ("Statement files processed", stats.get("statement_files_processed", "—")),
    ]:
        ws.cell(r, 2).value = label
        ws.cell(r, 2).font = Font(name="Calibri", bold=True, size=9, color="375623")
        ws.cell(r, 2).fill = LIGHT_FILL
        ws.cell(r, 2).alignment = MID_ALIGN
        ws.cell(r, 3).value = str(value)
        ws.cell(r, 3).font = BODY_FONT
        ws.cell(r, 3).fill = LIGHT_FILL
        ws.cell(r, 3).alignment = MID_ALIGN
        ws.row_dimensions[r].height = 20
        r += 1
    r += 1

    # Source files used
    source_files = stats.get("source_files", [])
    if source_files:
        ws.merge_cells(f"B{r}:D{r}")
        c = ws.cell(r, 2)
        c.value = "SOURCE FILES DETECTED"
        c.font = H1_FONT
        c.fill = H1_FILL
        c.alignment = MID_ALIGN
        ws.row_dimensions[r].height = 24
        r += 1
        for file_type, filename in source_files:
            ws.cell(r, 2).value = file_type
            ws.cell(r, 2).font = Font(name="Courier New", size=9, color="375623", bold=True)
            ws.cell(r, 2).fill = LIGHT_FILL
            ws.cell(r, 2).alignment = MID_ALIGN
            ws.merge_cells(f"C{r}:D{r}")
            ws.cell(r, 3).value = filename
            ws.cell(r, 3).font = BODY_FONT
            ws.cell(r, 3).fill = LIGHT_FILL
            ws.cell(r, 3).alignment = MID_ALIGN
            ws.row_dimensions[r].height = 20
            r += 1
        r += 1

    # Column source table
    SOURCE_MAP = [
        ("ISRC", "Products_Tracks export", "isrc"),
        ("Track Title", "Products_Tracks export", "track-title"),
        ("Track Version", "(not exported — populate manually)", "—"),
        ("UPC / Barcode", "Products_Tracks export", "upc"),
        ("Release Title", "Products_Tracks export", "album-title"),
        ("Cat No", "Products_Tracks / Contracts_wAlbumsTracks", "catalog-no"),
        ("Label", "Products_Tracks export", "label-name"),
        ("P Line", "Products_Tracks export", "p-line"),
        ("Release Date", "Products_Tracks export", "release-date"),
        ("Duration", "Products_Tracks export", "track-minutes + track-seconds"),
        ("Format", "Products_Tracks export", "product-type"),
        ("Distribution Channel", "Statement Sales CSVs", "Distribution Channel"),
        ("Price Category", "Contracts_Terms export", "price"),
        ("Catalogue Group", "(not exported — populate manually)", "—"),
        ("Dealer Price", "Products_Tracks export", "full-price (usd)"),
        ("Artist Name", "Contracts_wAlbumsTracks export", "artist"),
        ("Track Sales Contract Name", "Contracts_wAlbumsTracks export", "contract-title"),
        ("Track Sales Contract %", "Contracts_wAlbumsTracks export", "proration"),
        ("Track Costs Contract Name", "Contracts_wAlbumsTracks export", "contract-title (costs)"),
        ("Track Costs Contract %", "Contracts_wAlbumsTracks export", "proration (costs)"),
        ("Contract Name", "Contracts_Terms export", "contract-title"),
        ("Payee Name", "Contracts_wAlbumsTracks / Contracts_Terms", "payee"),
        ("Contract Type", "Contracts_Terms export", "rate-type"),
        ("Accounting Period", "(not exported — populate manually)", "—"),
        ("Currency", "Derived from statement data", "Original Currency"),
        ("Profit Share %", "Contracts_Terms export", "rate (when Profit Share)"),
        ("Contract Start Date", "Contracts_Terms export", "term-start"),
        ("Contract End Date", "Contracts_Terms export", "term-end"),
        ("Notes", "Contracts_Terms export", "comments"),
        ("Rate Territory", "Contracts_Terms export", "region"),
        ("Rate Channel", "Contracts_Terms export", "channel"),
        ("Rate Configuration", "Derived from price field", "price"),
        ("Rate Price Category", "Contracts_Terms export", "price"),
        ("Rate Source", "Contracts_Terms export", "source"),
        ("Rate Type / Basis", "Contracts_Terms export", "rate-type"),
        ("Rate %", "Contracts_Terms export", "rate"),
        ("Reserve %", "Contracts_Terms export", "reserve-rate"),
        ("Reserve Basis", "(not exported — populate manually)", "—"),
        ("Esc. Threshold Type", "(not exported separately)", "—"),
        ("Esc. When", "(not exported separately)", "—"),
        ("Esc. Threshold", "(not exported separately)", "—"),
        ("Esc. Rate %", "(not exported separately)", "—"),
        ("Balance (Latest Period)", "Statement Sales CSVs or Statement XLSXs", "Net Payable / Ending Balance:"),
        ("Advance Remaining", "(populate from contract or statement)", "—"),
        ("Min Payout", "(not exported)", "—"),
        ("Deduct Withholding Tax", "(not exported — populate manually)", "—"),
        ("Cross Contract Name", "Contracts_wAlbumsTracks export", "cross-collateralize"),
        ("Cross Contract Process", "(not fully exported)", "—"),
        ("Cross Contract Link", "(not fully exported)", "—"),
        ("Cross Contract %", "(not fully exported)", "—"),
    ]

    ws.merge_cells(f"B{r}:D{r}")
    c = ws.cell(r, 2)
    c.value = "COLUMN SOURCE TABLE"
    c.font = H1_FONT
    c.fill = H1_FILL
    c.alignment = MID_ALIGN
    ws.row_dimensions[r].height = 24
    r += 1

    for ci, lbl in [(2, "Column"), (3, "Source File"), (4, "Original Field Name")]:
        c = ws.cell(r, ci)
        c.value = lbl
        c.font = Font(name="Calibri", bold=True, size=9, color="666666")
        c.fill = PatternFill("solid", fgColor="E0EDE5")
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[r].height = 18
    r += 1

    alt = True
    for col_name, source, field in SOURCE_MAP:
        row_fill = LIGHT_FILL if alt else PatternFill("solid", fgColor="FFFFFF")
        col_band = next(
            (b for (b, h, *_) in COLUMN_DEFS if h == col_name), "recording"
        )
        ws.cell(r, 2).value = col_name
        ws.cell(r, 2).font = Font(
            name="Courier New", size=9, color=BANDS[col_band]["header"], bold=True
        )
        ws.cell(r, 2).fill = row_fill
        ws.cell(r, 2).alignment = MID_ALIGN

        ws.cell(r, 3).value = source
        ws.cell(r, 3).font = BODY_FONT
        ws.cell(r, 3).fill = row_fill
        ws.cell(r, 3).alignment = MID_ALIGN

        ws.cell(r, 4).value = field
        ws.cell(r, 4).font = Font(name="Courier New", size=9, color="375623")
        ws.cell(r, 4).fill = row_fill
        ws.cell(r, 4).alignment = MID_ALIGN

        ws.row_dimensions[r].height = 20
        r += 1
        alt = not alt


# ---------------------------------------------------------------------------
# Public API — state-based incremental model
# ---------------------------------------------------------------------------

def empty_state() -> dict:
    """Return a blank mapping state (nothing processed yet)."""
    return {
        "catalog": {},
        "contract_terms": {},
        "links": [],
        "net_by": {},            # {contract: {isrc: {period: float}}}
        "best_balance": {},      # {contract_lower: {balance, period, artist}}
        "orchard_contracts": {}, # {artist_lower: {type, contract_date, ...}}
        "advance_balances": {},  # {name_lower: {name, balance}}
        "source_files": [],      # [[type, filename], ...]
        "stmt_files_processed": 0,
    }


def apply_files_to_state(
    state: dict,
    files: dict,
    source_filenames: dict | None = None,
) -> dict:
    """
    Apply a batch of detected files to an existing mapping state.
    Catalog/contract/links uploads overwrite (they're snapshots).
    Statement uploads accumulate on top of existing balance data.
    Mutates state in place and returns it.
    """
    sfnames = source_filenames or {}

    # --- Snapshot files: overwrite on each upload ---
    if "catalog" in files:
        state["catalog"] = load_catalog(files["catalog"])
    if "contract_terms" in files:
        state["contract_terms"] = load_contract_terms(files["contract_terms"])
    if "isrc_links" in files:
        state["links"] = load_isrc_links(files["isrc_links"])
    if "orchard_contracts" in files:
        state["orchard_contracts"] = load_orchard_contracts(files["orchard_contracts"])
    if "advance_balances" in files:
        state["advance_balances"] = load_advance_balances(files["advance_balances"])

    # --- Statement files: accumulate ---
    # Reconstruct mutable defaultdicts from stored plain dicts
    net_by: dict = defaultdict(lambda: defaultdict(lambda: defaultdict(float)))
    for contract, isrc_dict in state.get("net_by", {}).items():
        for isrc, period_dict in isrc_dict.items():
            for period, val in period_dict.items():
                net_by[contract][isrc][period] += val

    best_balance: dict = dict(state.get("best_balance", {}))
    stmt_count = 0

    stmt_zips = files.get("statement_zip", [])
    if isinstance(stmt_zips, (bytes, bytearray)):
        stmt_zips = [stmt_zips]
    for zip_data in stmt_zips:
        stmt_count += load_statement_zip(zip_data, net_by, best_balance)

    stmt_csvs = files.get("statement_csv", [])
    if isinstance(stmt_csvs, (bytes, bytearray)):
        stmt_csvs = [stmt_csvs]
    for csv_data in stmt_csvs:
        if load_statement_csv(csv_data, net_by):
            stmt_count += 1

    stmt_xlsxs = files.get("statement_xlsx", [])
    if isinstance(stmt_xlsxs, (bytes, bytearray)):
        stmt_xlsxs = [stmt_xlsxs]
    xlsx_fnames = sfnames.get("statement_xlsx", [])
    if isinstance(xlsx_fnames, str):
        xlsx_fnames = [xlsx_fnames]
    for i, xlsx_data in enumerate(stmt_xlsxs):
        fname = xlsx_fnames[i] if i < len(xlsx_fnames) else ""
        if load_statement_xlsx(xlsx_data, fname, best_balance):
            stmt_count += 1

    # Persist back to plain dicts for JSON serialisation
    state["net_by"] = {
        contract: {isrc: dict(pd) for isrc, pd in id_.items()}
        for contract, id_ in net_by.items()
    }
    state["best_balance"] = best_balance
    state["stmt_files_processed"] = state.get("stmt_files_processed", 0) + stmt_count

    # Track source files (deduplicate by filename)
    existing_fnames = {entry[1] for entry in state.get("source_files", [])}
    for ftype, fname in sfnames.items():
        fnames = fname if isinstance(fname, list) else [fname]
        for fn in fnames:
            if fn not in existing_fnames:
                state.setdefault("source_files", []).append([ftype, fn])
                existing_fnames.add(fn)

    return state


def render_state_to_xlsx(state: dict) -> bytes:
    """Render the current mapping state to an XLSX workbook."""
    net_by: dict = defaultdict(lambda: defaultdict(lambda: defaultdict(float)))
    for contract, isrc_dict in state.get("net_by", {}).items():
        for isrc, period_dict in isrc_dict.items():
            for period, val in period_dict.items():
                net_by[contract][isrc][period] += val

    rows = _build_data_rows(
        state.get("catalog", {}),
        state.get("contract_terms", {}),
        state.get("links", []),
        net_by,
        state.get("best_balance", {}),
        orchard_contracts=state.get("orchard_contracts", {}),
        advance_balances=state.get("advance_balances", {}),
    )

    stats = {
        "row_count": len(rows),
        "isrc_count": len({r[0] for r in rows}),
        "artists_with_balance": len(state.get("best_balance", {})),
        "statement_files_processed": state.get("stmt_files_processed", 0),
        "source_files": state.get("source_files", []),
    }

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Mapping Template"
    ws.sheet_properties.tabColor = "B5696A"
    _build_mapping_sheet(ws, rows)

    wi = wb.create_sheet("Instructions")
    wi.sheet_properties.tabColor = "8B7BAB"
    _build_instructions_sheet(wi)

    wg = wb.create_sheet("Source Guide")
    wg.sheet_properties.tabColor = "6A9178"
    _build_source_guide_sheet(wg, stats)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def state_summary(state: dict) -> dict:
    """Return lightweight status info without rendering the XLSX."""
    net_by = state.get("net_by", {})
    return {
        "has_catalog": bool(state.get("catalog")),
        "has_links": bool(state.get("links")),
        "has_contract_terms": bool(state.get("contract_terms")),
        "has_orchard_contracts": bool(state.get("orchard_contracts")),
        "has_advance_balances": bool(state.get("advance_balances")),
        "isrc_count": len(state.get("catalog", {})),
        "contract_count": len(state.get("contract_terms", {})),
        "orchard_artist_count": len(state.get("orchard_contracts", {})),
        "stmt_files_processed": state.get("stmt_files_processed", 0),
        "contracts_with_balance": len(net_by) + len(state.get("best_balance", {})),
        "source_files": state.get("source_files", []),
    }


def build_blank_template() -> bytes:
    """XLSX with column structure + Instructions tab, no data rows."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Mapping Template"
    ws.sheet_properties.tabColor = "B5696A"
    _build_mapping_sheet(ws, [])

    wi = wb.create_sheet("Instructions")
    wi.sheet_properties.tabColor = "8B7BAB"
    _build_instructions_sheet(wi)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
