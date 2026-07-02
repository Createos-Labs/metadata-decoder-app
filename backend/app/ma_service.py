"""Service layer for M&A Audit scans and acquisitions."""
from __future__ import annotations

import json
import tempfile
import uuid
from datetime import datetime, timezone
from pathlib import Path

import engine  # noqa: F401 — puts engine/ on sys.path
import scan_ma as sm
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .db import get_db
from .storage import get_storage


def _now() -> str:
    return datetime.now(timezone.utc).isoformat()


def _key(scan_id: str, name: str) -> str:
    return f"ma/{scan_id}/{name}"


SEVERITIES = ("BLOCKER", "HIGH", "MEDIUM", "LOW", "INFO")

SEVERITY_COLORS = {
    "BLOCKER": "C00000",   # dark red
    "HIGH":    "FF0000",   # red
    "MEDIUM":  "FF9900",   # orange
    "LOW":     "FFD966",   # yellow
    "INFO":    "9DC3E6",   # light blue
}

HEADER_FILL = PatternFill("solid", start_color="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial")


def _autosize(ws, min_w=12, max_w=70):
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        longest = max(
            (len(str(c.value)) for c in col_cells if c.value is not None), default=min_w
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = max(min_w, min(longest + 2, max_w))


def _write_report(
    output_path: Path,
    acquisition: dict,
    scans_with_findings: list[dict],
) -> None:
    """
    Write a formatted Excel report.
    One tab = SUMMARY, then one tab per uploaded file.
    Only reviewed (non-dismissed, severity-assigned) findings are included.
    """
    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    # ---- SUMMARY tab -------------------------------------------------------
    ws_sum = wb.create_sheet("SUMMARY")
    ws_sum.append([f"{acquisition['name'].upper()} — DATA GAPS SUMMARY"])
    ws_sum["A1"].font = Font(bold=True, size=14, name="Arial")
    ws_sum.append([f"Prepared by: {acquisition.get('created_by', '')}  |  {datetime.now().strftime('%B %Y')}"])
    ws_sum.append([])

    # Severity legend
    ws_sum.append(["SEVERITY LEGEND"])
    ws_sum["A4"].font = Font(bold=True, name="Arial")
    legend = [
        ("BLOCKER", "Must be resolved before any payment can be made legally or operationally"),
        ("HIGH",    "Required for LE setup or first statement run — resolve before close activities"),
        ("MEDIUM",  "Important but not an immediate Day 1 stopper — resolve within 30 days"),
        ("LOW",     "Data hygiene / nice to have — address in Phase 2"),
        ("INFO",    "Informational — no immediate action required"),
    ]
    for sev, desc in legend:
        row_idx = ws_sum.max_row + 1
        ws_sum.append([sev, desc])
        color = SEVERITY_COLORS.get(sev, "FFFFFF")
        ws_sum.cell(row=row_idx, column=1).fill = PatternFill("solid", start_color=color)
        ws_sum.cell(row=row_idx, column=1).font = Font(bold=True, color="FFFFFF" if sev in ("BLOCKER", "HIGH") else "000000", name="Arial")

    ws_sum.append([])
    # Per-file counts
    count_header_row = ws_sum.max_row + 1
    ws_sum.append(["File"] + list(SEVERITIES))
    for cell in ws_sum[count_header_row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    totals = {s: 0 for s in SEVERITIES}
    for item in scans_with_findings:
        counts = {s: 0 for s in SEVERITIES}
        for f in item["findings"]:
            if not f.get("dismissed") and f.get("severity") in SEVERITIES:
                counts[f["severity"]] += 1
                totals[f["severity"]] += 1
        ws_sum.append([item["filename"]] + [counts[s] or "" for s in SEVERITIES])

    total_row = ws_sum.max_row + 1
    ws_sum.append(["TOTAL"] + [totals[s] or "" for s in SEVERITIES])
    ws_sum[f"A{total_row}"].font = Font(bold=True, name="Arial")

    _autosize(ws_sum)

    # ---- Per-file tabs -----------------------------------------------------
    for item in scans_with_findings:
        tab_name = item["filename"][:31]  # Excel sheet name limit
        ws = wb.create_sheet(tab_name)

        headers = ["Field / Issue", "Severity", "Finding", "Why It Matters", "Action Required"]
        ws.append(headers)
        for cell in ws[1]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.freeze_panes = "A2"

        reviewed = [
            f for f in item["findings"]
            if not f.get("dismissed") and f.get("severity") in SEVERITIES
        ]
        reviewed.sort(key=lambda f: (SEVERITIES.index(f["severity"]), f["sheet"], f["field"]))

        for finding in reviewed:
            sev = finding["severity"]
            color = SEVERITY_COLORS.get(sev, "FFFFFF")
            row_idx = ws.max_row + 1
            ws.append([
                finding["field"],
                sev,
                finding["finding"],
                finding["why_it_matters"],
                "",  # Action Required — left blank for analyst to fill
            ])
            sev_cell = ws.cell(row=row_idx, column=2)
            sev_cell.fill = PatternFill("solid", start_color=color)
            sev_cell.font = Font(
                bold=True,
                color="FFFFFF" if sev in ("BLOCKER", "HIGH") else "000000",
                name="Arial",
            )
            for col_idx in range(1, 6):
                ws.cell(row=row_idx, column=col_idx).alignment = Alignment(
                    wrap_text=True, vertical="top"
                )

        ws.row_dimensions[1].height = 20
        ws.column_dimensions["A"].width = 35
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 50
        ws.column_dimensions["D"].width = 55
        ws.column_dimensions["E"].width = 45

    wb.save(output_path)


class MAService:
    def __init__(self) -> None:
        self.storage = get_storage()
        self.db = get_db()

    # ---- Acquisitions -------------------------------------------------------

    def create_acquisition(self, name: str, company: str, user_email: str) -> dict:
        acq_id = uuid.uuid4().hex
        acq = {
            "id": acq_id,
            "name": name,
            "company": company,
            "status": "Active",
            "created_at": _now(),
            "updated_at": _now(),
            "created_by": user_email,
            "scan_ids": [],
        }
        self.db.create_acquisition(acq)
        return acq

    def list_acquisitions(self) -> list[dict]:
        return self.db.list_acquisitions()

    def get_acquisition(self, acq_id: str) -> dict | None:
        return self.db.get_acquisition(acq_id)

    def update_acquisition(self, acq_id: str, fields: dict) -> dict | None:
        allowed = {"name", "company", "status"}
        safe = {k: v for k, v in fields.items() if k in allowed}
        safe["updated_at"] = _now()
        self.db.update_acquisition(acq_id, safe)
        return self.db.get_acquisition(acq_id)

    def delete_acquisition(self, acq_id: str) -> bool:
        acq = self.db.get_acquisition(acq_id)
        if not acq:
            return False
        for scan_id in acq.get("scan_ids", []):
            self._delete_ma_scan(scan_id)
        return self.db.delete_acquisition(acq_id)

    # ---- MA Scans -----------------------------------------------------------

    def create_ma_scan(
        self, *, acq_id: str, data: bytes, filename: str, user_email: str
    ) -> dict:
        acq = self.db.get_acquisition(acq_id)
        if not acq:
            raise ValueError(f"Acquisition {acq_id!r} not found.")

        scan_id = uuid.uuid4().hex
        self.storage.write(_key(scan_id, "original.xlsx"), data)

        with tempfile.TemporaryDirectory() as td:
            tmp = Path(td)
            in_path = tmp / "input.xlsx"
            in_path.write_bytes(data)
            result = sm.analyze(in_path)

        findings = result["findings"]
        stats = result["stats"]

        self.storage.write(
            _key(scan_id, "findings.json"),
            json.dumps(findings, ensure_ascii=False).encode(),
        )

        scan = {
            "id": scan_id,
            "acquisition_id": acq_id,
            "filename": filename,
            "created_at": _now(),
            "updated_at": _now(),
            "uploaded_by": user_email,
            "sheets_scanned": result["sheets_scanned"],
            "total_findings": stats["total_findings"],
            "reviewed_count": 0,
            "dismissed_count": 0,
            "sheet_stats": stats["sheets"],
        }
        self.db.create_ma_scan(scan)

        # Register this scan in the acquisition.
        updated_ids = acq.get("scan_ids", []) + [scan_id]
        self.db.update_acquisition(acq_id, {
            "scan_ids": updated_ids,
            "updated_at": _now(),
        })

        return {"scan": scan, "findings": findings}

    def get_ma_scan(self, scan_id: str) -> dict | None:
        return self.db.get_ma_scan(scan_id)

    def get_findings(self, scan_id: str) -> list[dict] | None:
        raw = self.storage.read(_key(scan_id, "findings.json"))
        if raw is None:
            return None
        return json.loads(raw)

    def list_scans_for_acquisition(self, acq_id: str) -> list[dict]:
        return self.db.list_ma_scans_for_acquisition(acq_id)

    def update_findings(self, scan_id: str, updates: list[dict]) -> list[dict]:
        """
        Apply severity assignments and dismissals from the analyst.
        updates = [{"id": "...", "severity": "HIGH"|null, "dismissed": bool}, ...]
        Carries forward previous assignments for findings not in the update list.
        """
        findings = self.get_findings(scan_id) or []
        update_map = {u["id"]: u for u in updates}

        reviewed = 0
        dismissed = 0
        for f in findings:
            if f["id"] in update_map:
                u = update_map[f["id"]]
                f["severity"] = u.get("severity")
                f["dismissed"] = bool(u.get("dismissed", False))
            if f.get("dismissed"):
                dismissed += 1
            elif f.get("severity") in SEVERITIES:
                reviewed += 1

        self.storage.write(
            _key(scan_id, "findings.json"),
            json.dumps(findings, ensure_ascii=False).encode(),
        )
        self.db.update_ma_scan(scan_id, {
            "updated_at": _now(),
            "reviewed_count": reviewed,
            "dismissed_count": dismissed,
        })
        return findings

    def _delete_ma_scan(self, scan_id: str) -> None:
        self.storage.delete_prefix(f"ma/{scan_id}/")
        self.db.delete_ma_scan(scan_id)

    def delete_ma_scan(self, scan_id: str, acq_id: str) -> bool:
        scan = self.db.get_ma_scan(scan_id)
        if not scan:
            return False
        self._delete_ma_scan(scan_id)
        # Remove from acquisition's scan list.
        acq = self.db.get_acquisition(acq_id)
        if acq:
            updated = [s for s in acq.get("scan_ids", []) if s != scan_id]
            self.db.update_acquisition(acq_id, {"scan_ids": updated, "updated_at": _now()})
        return True

    # ---- Report generation --------------------------------------------------

    def generate_report(self, acq_id: str) -> bytes | None:
        acq = self.db.get_acquisition(acq_id)
        if not acq:
            return None
        scans = self.db.list_ma_scans_for_acquisition(acq_id)
        scans_with_findings = []
        for scan in scans:
            findings = self.get_findings(scan["id"]) or []
            scans_with_findings.append({"filename": scan["filename"], "findings": findings})

        with tempfile.TemporaryDirectory() as td:
            out = Path(td) / "report.xlsx"
            _write_report(out, acq, scans_with_findings)
            return out.read_bytes()


_service: MAService | None = None


def get_ma_service() -> MAService:
    global _service
    if _service is None:
        _service = MAService()
    return _service
